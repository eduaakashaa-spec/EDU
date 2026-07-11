"""
Membership blueprint — replaces the Apps Script backend.

Public:
    POST /membership/apply                 ← was doPost / handleApplicationSubmit

Admin (login required + admin tier):
    GET  /admin/membership                 ← list everything (search/filter/paginate)
    GET  /admin/membership/<id>            ← single application detail
    POST /admin/membership/<id>/update     ← edit fields / status / notes
    POST /admin/membership/<id>/discount   ← set discount + reason
    POST /admin/membership/<id>/adhoc      ← set ad-hoc items
    POST /admin/membership/<id>/invoice    ← issue invoice (transaction-safe number)
    POST /admin/membership/<id>/payment    ← record a payment / issue receipt
    GET  /admin/membership/export.xlsx     ← download everything as Excel

PDF generation and email sending are Phase B (stubs are clearly marked).
"""
import io
import json
from datetime import datetime, timezone

from flask import (Blueprint, request, jsonify, render_template, redirect,
                   url_for, flash, abort, current_app, send_file)
from flask_login import current_user

from app.extensions import db
from app.decorators import admin_required
from app.models_membership import (MembershipApplication, MembershipInvoice,
                                   DocSequence, TIERS, APP_STATUSES)
from app.services.pricing import compute_totals, rupees, to_paise
from app.services.queries import count_if

membership_bp = Blueprint('membership', __name__)


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _seq(name, prefix):
    """Fetch (or lazily create) a DocSequence row, locked for update."""
    seq = db.session.get(DocSequence, name)
    if seq is None:
        seq = DocSequence(name=name, prefix=prefix, value=0)
        db.session.add(seq)
        db.session.flush()
    return seq


def _allocate(name, prefix, width=4):
    """Allocate the next number in a sequence, transaction-safe."""
    # Row-level lock so two concurrent requests can't get the same number.
    seq = (db.session.query(DocSequence)
           .filter_by(name=name)
           .with_for_update()
           .first())
    if seq is None:
        seq = _seq(name, prefix)
    seq.value += 1
    return f'{seq.prefix}{str(seq.value).zfill(width)}'


def _recalc(app_row):
    """Recompute the GST breakdown for an application from its current inputs."""
    t = compute_totals(
        tier_price=app_row.tier_price,
        addon_amount=app_row.addon_amount,
        adhoc_amount=app_row.adhoc_amount,
        discount=app_row.discount,
        customer_state=app_row.state or '',
        home_state=current_app.config['HOME_STATE'],
        gst_rate=current_app.config['GST_RATE'],
    )
    app_row.taxable_value = t.taxable_value
    app_row.cgst = t.cgst
    app_row.sgst = t.sgst
    app_row.igst = t.igst
    app_row.total_gst = t.total_gst
    app_row.final_total = t.final_total
    return t


# --------------------------------------------------------------------------- #
# public intake  (replaces Apps Script doPost)
# --------------------------------------------------------------------------- #
@membership_bp.route('/membership/apply', methods=['POST'])
def apply():
    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()

    if not name or not email:
        return jsonify({'ok': False, 'error': 'Name and email are required.'}), 400

    app_row = MembershipApplication(
        name=name,
        email=email,
        phone=(data.get('phone') or '').strip(),
        city=(data.get('city') or '').strip(),
        state=(data.get('state') or '').strip(),
        tier=(data.get('tier') or '').strip(),
        status='New',
    )
    # Allocate a human reference like EA-PREM-0001 (transaction-safe).
    app_row.reference = _allocate('reference', 'EA-PREM-')
    db.session.add(app_row)
    db.session.commit()

    # Phase B: email_application_received(app_row)
    return jsonify({'ok': True, 'reference': app_row.reference})


# --------------------------------------------------------------------------- #
# admin — list / search / filter  (the spreadsheet replacement)
# --------------------------------------------------------------------------- #
@membership_bp.route('/admin/membership')
@admin_required
def admin_list():
    q = (request.args.get('q') or '').strip()
    status = (request.args.get('status') or '').strip()
    tier = (request.args.get('tier') or '').strip()
    page = request.args.get('page', 1, type=int)

    query = MembershipApplication.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            MembershipApplication.name.ilike(like),
            MembershipApplication.email.ilike(like),
            MembershipApplication.phone.ilike(like),
            MembershipApplication.reference.ilike(like),
        ))
    if status:
        query = query.filter(MembershipApplication.status == status)
    if tier:
        query = query.filter(MembershipApplication.tier == tier)

    query = query.order_by(MembershipApplication.created_at.desc())
    pagination = query.paginate(page=page, per_page=25, error_out=False)

    # KPI summary cards — one aggregate SELECT instead of four COUNTs
    total, new, invoiced, paid = db.session.query(
        db.func.count(MembershipApplication.id),
        count_if(MembershipApplication.status == 'New'),
        count_if(MembershipApplication.status == 'Invoiced'),
        count_if(MembershipApplication.status == 'Paid')).one()
    stats = {'total': total, 'new': new, 'invoiced': invoiced, 'paid': paid}

    return render_template('admin/membership_list.html',
                           pagination=pagination, rows=pagination.items,
                           stats=stats, rupees=rupees,
                           q=q, status=status, tier=tier,
                           TIERS=TIERS, STATUSES=APP_STATUSES)


@membership_bp.route('/admin/membership/<int:app_id>')
@admin_required
def admin_detail(app_id):
    app_row = db.session.get(MembershipApplication, app_id) or abort(404)
    return render_template('admin/membership_detail.html',
                           a=app_row, rupees=rupees,
                           TIERS=TIERS, STATUSES=APP_STATUSES)


# --------------------------------------------------------------------------- #
# admin — mutations
# --------------------------------------------------------------------------- #
@membership_bp.route('/admin/membership/<int:app_id>/update', methods=['POST'])
@admin_required
def admin_update(app_id):
    a = db.session.get(MembershipApplication, app_id) or abort(404)
    f = request.form
    a.name = (f.get('name') or a.name).strip()
    a.phone = (f.get('phone') or '').strip()
    a.city = (f.get('city') or '').strip()
    a.state = (f.get('state') or '').strip()
    a.tier = (f.get('tier') or '').strip()
    if f.get('status') in APP_STATUSES:
        a.status = f.get('status')
    a.internal_notes = (f.get('internal_notes') or '').strip()
    if f.get('tier_price') is not None:
        a.tier_price = to_paise(f.get('tier_price'))
    if f.get('addon_amount') is not None:
        a.addon_amount = to_paise(f.get('addon_amount'))
    _recalc(a)
    db.session.commit()
    flash('Application updated.', 'success')
    return redirect(url_for('membership.admin_detail', app_id=a.id))


@membership_bp.route('/admin/membership/<int:app_id>/discount', methods=['POST'])
@admin_required
def admin_discount(app_id):
    a = db.session.get(MembershipApplication, app_id) or abort(404)
    a.discount = to_paise(request.form.get('discount'))
    a.discount_reason = (request.form.get('discount_reason') or '').strip()
    _recalc(a)
    db.session.commit()
    flash('Discount applied.', 'success')
    return redirect(url_for('membership.admin_detail', app_id=a.id))


@membership_bp.route('/admin/membership/<int:app_id>/adhoc', methods=['POST'])
@admin_required
def admin_adhoc(app_id):
    a = db.session.get(MembershipApplication, app_id) or abort(404)
    raw = request.form.get('adhoc_items_json') or '[]'
    try:
        items = json.loads(raw)
        total = sum(to_paise(it.get('amount', 0)) for it in items)
    except (ValueError, AttributeError):
        flash('Ad-hoc items must be valid JSON like [{"label":"X","amount":500}].', 'danger')
        return redirect(url_for('membership.admin_detail', app_id=a.id))
    a.adhoc_items_json = raw
    a.adhoc_amount = total
    _recalc(a)
    db.session.commit()
    flash('Ad-hoc items saved.', 'success')
    return redirect(url_for('membership.admin_detail', app_id=a.id))


@membership_bp.route('/admin/membership/<int:app_id>/invoice', methods=['POST'])
@admin_required
def admin_invoice(app_id):
    a = db.session.get(MembershipApplication, app_id) or abort(404)
    _recalc(a)
    inv = a.invoice
    if inv is None:
        inv = MembershipInvoice(application_id=a.id)
        db.session.add(inv)
    if not inv.invoice_no:
        inv.invoice_no = _allocate('invoice', 'EA/')
    inv.invoice_date = datetime.now(timezone.utc)
    inv.balance_due = a.final_total - (inv.amount_paid or 0)
    inv.payment_status = 'Unpaid' if (inv.amount_paid or 0) == 0 else inv.payment_status
    a.status = 'Invoiced'
    db.session.commit()
    # Phase B: build_document_pdf(inv, kind='invoice') + email_invoice(a, inv)
    flash(f'Invoice {inv.invoice_no} issued. (PDF + email: Phase B)', 'success')
    return redirect(url_for('membership.admin_detail', app_id=a.id))


@membership_bp.route('/admin/membership/<int:app_id>/payment', methods=['POST'])
@admin_required
def admin_payment(app_id):
    a = db.session.get(MembershipApplication, app_id) or abort(404)
    inv = a.invoice
    if inv is None:
        flash('Issue an invoice before recording a payment.', 'danger')
        return redirect(url_for('membership.admin_detail', app_id=a.id))

    inv.amount_paid = (inv.amount_paid or 0) + to_paise(request.form.get('amount'))
    inv.payment_mode = (request.form.get('payment_mode') or '').strip()
    inv.payment_ref = (request.form.get('payment_ref') or '').strip()
    inv.payment_date = datetime.now(timezone.utc)
    inv.balance_due = max(a.final_total - inv.amount_paid, 0)

    if inv.amount_paid >= a.final_total and a.final_total > 0:
        inv.payment_status = 'Paid'
        a.status = 'Paid'
        if not inv.receipt_no:
            inv.receipt_no = _allocate('receipt', 'RCPT/')
        # Phase B: build_document_pdf(inv, kind='receipt') + email_receipt(a, inv)
    elif inv.amount_paid > 0:
        inv.payment_status = 'Partially Paid'
        a.status = 'Partially Paid'

    db.session.commit()
    flash('Payment recorded.', 'success')
    return redirect(url_for('membership.admin_detail', app_id=a.id))


# --------------------------------------------------------------------------- #
# admin — Excel export  (so non-tech staff still get their spreadsheet)
# --------------------------------------------------------------------------- #
@membership_bp.route('/admin/membership/export.xlsx')
@admin_required
def admin_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # Respect the same filters as the list view.
    q = (request.args.get('q') or '').strip()
    status = (request.args.get('status') or '').strip()
    tier = (request.args.get('tier') or '').strip()

    query = MembershipApplication.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            MembershipApplication.name.ilike(like),
            MembershipApplication.email.ilike(like),
            MembershipApplication.reference.ilike(like),
        ))
    if status:
        query = query.filter(MembershipApplication.status == status)
    if tier:
        query = query.filter(MembershipApplication.tier == tier)
    rows = query.order_by(MembershipApplication.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Requests'

    headers = ['Reference', 'Created', 'Status', 'Name', 'Email', 'Phone',
               'City', 'State', 'Tier', 'Tier Price', 'Add-on Amount',
               'Discount', 'Ad-hoc Amount', 'Taxable Value', 'CGST', 'SGST',
               'IGST', 'Total GST', 'Final Total', 'Invoice No', 'Invoice Date',
               'Payment Status', 'Amount Paid', 'Balance Due', 'Receipt No',
               'Internal Notes']
    ws.append(headers)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='0E3A8A')  # EduAakashaa navy
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill

    def r(p):  # paise -> rupees float for spreadsheet maths
        return round((p or 0) / 100, 2)

    for a in rows:
        inv = a.invoice
        ws.append([
            a.reference,
            a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else '',
            a.status, a.name, a.email, a.phone, a.city, a.state, a.tier,
            r(a.tier_price), r(a.addon_amount), r(a.discount), r(a.adhoc_amount),
            r(a.taxable_value), r(a.cgst), r(a.sgst), r(a.igst), r(a.total_gst),
            r(a.final_total),
            inv.invoice_no if inv else '',
            inv.invoice_date.strftime('%Y-%m-%d') if inv and inv.invoice_date else '',
            inv.payment_status if inv else '',
            r(inv.amount_paid) if inv else 0,
            r(inv.balance_due) if inv else 0,
            inv.receipt_no if inv else '',
            a.internal_notes or '',
        ])

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'eduaakashaa_memberships_{datetime.now():%Y%m%d}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# --------------------------------------------------------------------------- #
# admin — leads, contact inquiries, users (all captured data in one place)
# --------------------------------------------------------------------------- #
@membership_bp.route('/admin/leads')
@admin_required
def admin_leads():
    from app.models import PageLead, DasaLead

    q = (request.args.get('q') or '').strip()
    source = (request.args.get('source') or '').strip()
    page = request.args.get('page', 1, type=int)

    query = PageLead.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            PageLead.name.ilike(like),
            PageLead.email.ilike(like),
            PageLead.phone.ilike(like),
        ))
    if source:
        query = query.filter(PageLead.source == source)

    pagination = (query.order_by(PageLead.created_at.desc())
                  .paginate(page=page, per_page=25, error_out=False))

    # One GROUP BY yields both the source list and the grand total.
    src_counts = (db.session.query(PageLead.source, db.func.count(PageLead.id))
                  .group_by(PageLead.source).order_by(PageLead.source).all())
    sources = [s for s, _ in src_counts]
    stats = {
        'total': sum(c for _, c in src_counts),
        'sources': len(sources),
        'dasa_leads': db.session.query(db.func.count(DasaLead.id)).scalar(),
    }
    dasa_rows = (DasaLead.query.order_by(DasaLead.timestamp.desc()).limit(25).all()
                 if not q and not source else [])

    return render_template('admin/leads_list.html', admin_tab='leads',
                           pagination=pagination, rows=pagination.items,
                           sources=sources, stats=stats, q=q, source=source,
                           dasa_rows=dasa_rows)


@membership_bp.route('/admin/leads/<int:lead_id>')
@admin_required
def admin_lead_detail(lead_id):
    from app.models import PageLead
    lead = db.session.get(PageLead, lead_id) or abort(404)
    try:
        pretty = json.dumps(json.loads(lead.payload_json or '{}'),
                            indent=2, ensure_ascii=False)
    except ValueError:
        pretty = lead.payload_json or ''
    return render_template('admin/lead_detail.html', admin_tab='leads',
                           lead=lead, pretty=pretty)


@membership_bp.route('/admin/inquiries')
@admin_required
def admin_inquiries():
    from app.models import ContactInquiry

    q = (request.args.get('q') or '').strip()
    page = request.args.get('page', 1, type=int)

    query = ContactInquiry.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            ContactInquiry.first_name.ilike(like),
            ContactInquiry.last_name.ilike(like),
            ContactInquiry.email.ilike(like),
            ContactInquiry.message.ilike(like),
        ))

    pagination = (query.order_by(ContactInquiry.created_at.desc())
                  .paginate(page=page, per_page=25, error_out=False))
    return render_template('admin/inquiries_list.html', admin_tab='inquiries',
                           pagination=pagination, rows=pagination.items,
                           total=ContactInquiry.query.count(), q=q)


@membership_bp.route('/admin/users')
@admin_required
def admin_users():
    from app.models import User

    q = (request.args.get('q') or '').strip()
    tier_filter = (request.args.get('tier') or '').strip()
    page = request.args.get('page', 1, type=int)

    query = User.query
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(User.name.ilike(like), User.email.ilike(like)))
    if tier_filter in ('free', 'premium', 'admin'):
        query = query.filter(User.tier == tier_filter)

    pagination = (query.order_by(User.created_at.desc())
                  .paginate(page=page, per_page=25, error_out=False))
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    total, premium, admin, expired = db.session.query(
        db.func.count(User.id),
        count_if(User.tier == 'premium'),
        count_if(User.tier == 'admin'),
        count_if(db.and_(User.tier == 'premium',
                         User.tier_expires_at.isnot(None),
                         User.tier_expires_at <= now))).one()
    stats = {'total': total, 'premium': premium, 'admin': admin,
             'expired': expired}
    return render_template('admin/users_list.html', admin_tab='users',
                           pagination=pagination, rows=pagination.items,
                           stats=stats, q=q, tier_filter=tier_filter)
